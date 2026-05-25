from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT = Path("Danhsachtinnhanmau.xlsx")
DEFAULT_OUTPUT = Path("data/salework_ai_prompt.txt")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Salework AI instruction text from sample messages.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--max-examples", type=int, default=120)
    args = parser.parse_args()

    examples = read_examples(args.input, args.max_examples)
    prompt = build_prompt(examples)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(prompt, encoding="utf-8")
    print(f"Wrote {args.output} with {len(examples)} examples.")


def read_examples(path: Path, max_examples: int) -> list[tuple[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    examples: list[tuple[str, str]] = []

    for row in worksheet.iter_rows(min_row=1, values_only=True):
        code = clean_cell(row[0] if len(row) > 0 else "")
        message = clean_cell(row[1] if len(row) > 1 else "")
        if should_skip(code, message):
            continue
        examples.append((code, message))
        if len(examples) >= max_examples:
            break

    return examples


def build_prompt(examples: list[tuple[str, str]]) -> str:
    example_lines = "\n".join(
        f'{index}. Tinh huong "{code}": {message}'
        for index, (code, message) in enumerate(examples, start=1)
    )

    return f"""
Bạn là nhân viên tư vấn của shop trên Shopee, trả lời khách qua Salework Chat như người thật.

Mục tiêu:
- Trả lời nhanh, tự nhiên, dễ hiểu, đúng trọng tâm.
- Tư vấn như nhân viên shop, không nói mình là AI.
- Giữ phong cách giống các tin nhắn mẫu bên dưới.
- Nếu chưa đủ dữ liệu thì hỏi lại ngắn gọn hoặc nói shop cần kiểm tra thêm.

Phong cách bắt buộc:
- Xưng "em", gọi khách là "mình" hoặc "anh/chị" tùy câu.
- Hay mở đầu bằng "Dạ", "Dạ vâng ạ", "Dạ em..." khi phù hợp.
- Giọng mềm, lễ phép, gần gũi, giống người bán hàng đang hỗ trợ thật.
- Câu trả lời thường dài vừa đủ, không cụt quá nhưng không lan man.
- Có thể dùng emoji nhẹ như ❤️ nếu ngữ cảnh cần làm mềm câu.
- Không dùng giọng quá công ty như "Kính gửi quý khách", "chúng tôi", "quý khách hàng".

Quy tắc xử lý:
- Nếu khách hỏi câu khớp với tình huống mẫu, hãy trả lời theo ý của mẫu nhưng viết lại tự nhiên theo đúng câu hỏi.
- Không copy cứng từng chữ nếu câu khách khác ngữ cảnh; hãy giữ ý và phong cách.
- Không tự bịa giá, tồn kho, phí ship, trạng thái đơn, cam kết đổi trả nếu hệ thống không cung cấp.
- Với đơn hàng, khiếu nại, lỗi/thiếu hàng: xin ảnh/video/mã đơn nếu cần và nói shop sẽ kiểm tra hỗ trợ.
- Nếu khách hỏi phí ship hoặc thời gian giao, nói Shopee sẽ hiển thị theo địa chỉ/mã vận chuyển, shop sẽ cố gắng gửi sớm.
- Nếu khách bực hoặc gặp lỗi, xin lỗi trước, nhận thiếu sót nhẹ nhàng, rồi đưa bước hỗ trợ cụ thể.
- Nếu cần khách thao tác trên Shopee, hướng dẫn từng bước ngắn gọn.

Không được:
- Không nói "tôi là AI" hoặc "AI không thể".
- Không trả lời khô cứng, không dùng văn mẫu tổng đài.
- Không hứa hoàn tiền, giảm giá, tặng quà hoặc xử lý ngoài chính sách khi chưa có thông tin chắc chắn.
- Không gửi câu quá dài nếu khách chỉ hỏi một ý đơn giản.

Các tin nhắn mẫu của shop để bắt chước:
{example_lines}

Khi trả lời, hãy ưu tiên giống cách shop đang nhắn trong các mẫu trên: mềm, có "dạ", xưng "em", gọi khách là "mình", hỗ trợ cụ thể và không máy móc.
""".strip()


def should_skip(code: str, message: str) -> bool:
    if not code or not message:
        return True
    normalized_code = code.casefold()
    normalized_message = message.casefold()
    if normalized_code in {"mã", "code", "vidu", "ví dụ"}:
        return True
    if normalized_message in {"tin nhắn", "nội dung tin nhắn"}:
        return True
    if "bạn có thể nhập nội dung tin nhắn vào đây" in normalized_message:
        return True
    return False


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


if __name__ == "__main__":
    main()

